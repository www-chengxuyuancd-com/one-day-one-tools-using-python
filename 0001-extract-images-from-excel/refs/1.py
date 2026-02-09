import os
import re
import sys
import warnings
from pathlib import Path
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import openpyxl.utils
import win32com.client
import time
import json
import logging
import requests
from urllib.parse import urlparse
import colorama
from colorama import Fore, Back, Style
import tqdm
from datetime import datetime
from PIL import Image

# 初始化colorama
colorama.init()

# 过滤特定的警告信息
warnings.filterwarnings('ignore', category=UserWarning, 
                       module='openpyxl.reader.drawings')

def get_exe_dir():
    """获取程序所在目录"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe运行
        return Path(sys.executable).parent
    else:
        # 如果是源码运行
        return Path(__file__).parent

def save_last_config(config):
    """保存最后一次使用的配置到程序所在目录"""
    save_config = {
        'sheet_name': config['sheet_name'],
        'name_column': config['name_column'],
        'image_column': config['image_column'],
        'start_row': config['start_row'],
        'naming_method': config['naming_method'],
        'custom_prefix': config['custom_prefix'],
        'custom_suffix': config['custom_suffix'],
        'connector': config['connector'],
        'image_format': config['image_format']
    }
    try:
        config_path = get_exe_dir() / 'last_config.json'
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(save_config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print_status(f"保存配置文件失败: {str(e)}", "warning")

def load_last_config():
    """从程序所在目录加载上次使用的配置"""
    try:
        config_path = get_exe_dir() / 'last_config.json'
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print_status(f"加载配置文件失败: {str(e)}", "warning")
    return None

def get_config(excel_path, wb=None, retry_sheet=False):
    """根据Excel文件路径生成配置"""
    last_config = load_last_config()
    
    print("\n请输入配置信息（直接回车使用方括号中的值）：")
    
    default_sheet = last_config['sheet_name'] if last_config else '商品数据'
    default_name_col = last_config['name_column'] if last_config else 'AE'
    default_img_col = last_config['image_column'] if last_config else 'A'
    default_start_row = last_config['start_row'] if last_config else 2
    default_naming_method = last_config.get('naming_method', '1') if last_config else '1'
    default_custom_prefix = last_config.get('custom_prefix', '') if last_config else ''
    default_custom_suffix = last_config.get('custom_suffix', '') if last_config else ''
    default_connector = last_config.get('connector', '') if last_config else ''
    default_image_format = last_config.get('image_format', 'png') if last_config else 'png'
    
    # 显示工作表列表
    if wb:
        print_colored("\n当前Excel文件包含以下工作表：", Fore.YELLOW)
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print_colored(f"  {i}. {sheet_name}", Fore.CYAN)
        print_colored("  提示：可以输入序号或工作表名称", Fore.GREEN)
    
    # 获取工作表名称
    while True:
        sheet_input = input(f"请输入工作表序号或名称 [{default_sheet}]: ").strip()
        if not sheet_input:
            sheet_name = default_sheet
            break
        
        # 尝试解析为序号
        if sheet_input.isdigit():
            idx = int(sheet_input)
            if 1 <= idx <= len(wb.sheetnames):
                sheet_name = wb.sheetnames[idx-1]
                break
            else:
                print_status("无效的序号，请重新输入", "error")
                continue
        
        # 直接使用输入的名称
        if sheet_input in wb.sheetnames:
            sheet_name = sheet_input
            break
        else:
            print_status("工作表不存在，请重新输入", "error")
            continue
    
    name_column = input(f"请输入命名内容所在列 [{default_name_col}]: ").strip().upper()
    image_column = input(f"请输入图片所在列 [{default_img_col}]: ").strip().upper()
    start_row = input(f"请输入起始行号 [{default_start_row}]: ").strip()
    
    print("\n图片命名方式：")
    print("1. 使用命名内容命名")
    print("2. 使用自定义格式命名")
    naming_method = input(f"请选择命名方式 (1/2) [{default_naming_method}]: ").strip()
    naming_method = naming_method if naming_method in ['1', '2'] else default_naming_method
    
    # 显示支持的图片格式列表
    print("\n支持的图片格式：")
    formats = ['png', 'jpg', 'jpeg', 'webp', 'bmp', 'gif']
    for i, fmt in enumerate(formats, 1):
        print_colored(f"  {i}. {fmt}", Fore.CYAN)
    print_colored("  提示：可以输入序号或格式名称", Fore.GREEN)
    print_colored("  注意：某些格式可能不支持透明度", Fore.YELLOW)
    
    # 获取图片格式
    while True:
        format_input = input(f"请输入格式序号或名称 [{default_image_format}]: ").strip().lower()
        if not format_input:
            image_format = default_image_format
            break
        
        # 尝试解析为序号
        if format_input.isdigit():
            idx = int(format_input)
            if 1 <= idx <= len(formats):
                image_format = formats[idx-1]
                break
            else:
                print_status("无效的序号，请重新输入", "error")
                continue
        
        # 直接使用输入的格式
        if format_input in formats:
            image_format = format_input
            break
        else:
            print_status("不支持的格式，请重新输入", "error")
            continue
    
    custom_prefix = ''
    custom_suffix = ''
    connector = ''
    if naming_method == '2':
        print("\n自定义格式配置：")
        print("最终格式将为：前缀+连接符+后缀")
        print("特殊用法：如果只设置后缀（前缀和连接符留空），将使用从1开始的递增数字命名")
        print("例如：只设置后缀时，文件将依次命名为：1.png、2.png、3.png...")
        print("普通示例：前缀为'img'，连接符为'_'，后缀为'001'时，将命名为：img_001.png")
        
        custom_prefix = input(f"请输入前缀 [{default_custom_prefix}]: ").strip()
        custom_prefix = custom_prefix if custom_prefix else default_custom_prefix
        
        connector = input(f"请输入连接符 [{default_connector}]: ").strip()
        connector = connector if connector else default_connector
        
        custom_suffix = input(f"请输入后缀 [{default_custom_suffix}]: ").strip()
        custom_suffix = custom_suffix if custom_suffix else default_custom_suffix
    
    config = {
        'base_path': str(excel_path.parent),
        'excel_filename': excel_path.name,
        'sheet_name': sheet_name if sheet_name else default_sheet,
        'name_column': name_column if name_column else default_name_col,
        'image_column': image_column if image_column else default_img_col,
        'start_row': int(start_row) if start_row.isdigit() else default_start_row,
        'naming_method': naming_method,
        'custom_prefix': custom_prefix,
        'custom_suffix': custom_suffix,
        'connector': connector,
        'image_format': image_format
    }
    
    save_last_config(config)
    return config

def get_image_safely(loader, cell_ref):
    """使用 SheetImageLoader 获取图片"""
    try:
        if loader.image_in(cell_ref):
            return loader.get(cell_ref)
    except Exception as e:
        print_colored(f"图片提取警告 [{cell_ref}]: {str(e)}", Fore.YELLOW)
    return None

def ensure_excel_anchors(excel_path):
    """使用WPS打开并保存文件，确保图片锚点信息完整"""
    wps = None
    try:
        excel_path = str(excel_path.resolve())  # 获取完整路径
        
        # 尝试不同的应用程序，一旦成功就停止尝试
        apps = [
            ("WPS.Application", "WPS"),
            ("KET.Application", "KET"),
            ("Excel.Application", "Excel")
        ]
        
        for app_name, app_type in apps:
            try:
                wps = win32com.client.Dispatch(app_name)
                print_status(f"使用 {app_type} 进行预处理...", "info")
                break
            except:
                continue
                
        if not wps:
            print_status("无法启动办公软件进行预处理，尝试直接处理...", "warning")
            return False
        
        wps.Visible = False
        wps.DisplayAlerts = False
        
        # 打开工作簿并立即保存关闭
        workbook = wps.Workbooks.Open(excel_path)
        time.sleep(0.1)  # 极短暂等待
        workbook.Save()
        workbook.Close()
        
        return True
        
    except Exception as e:
        print_status(f"文件预处理警告: {str(e)}", "warning")
        return False
    finally:
        if wps:
            try:
                wps.Quit()
                time.sleep(0.1)  # 极短暂等待
            except:
                pass

def validate_config(wb, config):
    """验证配置是否有效"""
    errors = []
    
    # 验证工作表名称
    if config['sheet_name'] not in wb.sheetnames:
        errors.append(f"工作表'{config['sheet_name']}'不存在")
        print(f"可用的工作表: {', '.join(wb.sheetnames)}")
    
    # 验证列名格式
    column_pattern = re.compile(r'^[A-Z]+$')
    if not column_pattern.match(config['name_column']):
        errors.append(f"命名内容列'{config['name_column']}'格式无效")
    if not column_pattern.match(config['image_column']):
        errors.append(f"图片列'{config['image_column']}'格式无效")
    
    # 验证起始行
    if config['start_row'] < 1:
        errors.append("起始行号必须大于0")
    
    return errors

def print_progress(current, total, processed, failed):
    """显示进度条"""
    width = 50
    progress = current / total
    filled = int(width * progress)
    bar = '=' * filled + '-' * (width - filled)
    percent = progress * 100
    print(f'\r处理进度: [{bar}] {percent:5.1f}% | 成功: {processed} | 失败: {failed}', end='')

def get_safe_filename(name, max_length=100):
    """生成安全的文件名"""
    # 移除非法字符
    name = re.sub(r'[\\/*?:"<>|]', "_", str(name).strip())
    # 限制长度
    if len(name) > max_length:
        name = name[:max_length-3] + "..."
    return name or "未命名"

def setup_logging(save_dir):
    """设置日志"""
    # 创建日志文件名（包含时间戳）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = save_dir / f"处理日志_{timestamp}.txt"
    
    # 配置日志格式
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def log_config(logger, config):
    """记录配置信息到日志"""
    logger.info("=" * 50)
    logger.info("配置信息:")
    logger.info("-" * 20)
    logger.info(f"工作表: {config['sheet_name']}")
    logger.info(f"命名内容列: {config['name_column']}")
    logger.info(f"图片列: {config['image_column']}")
    logger.info(f"起始行: {config['start_row']}")
    logger.info(f"图片格式: {config['image_format']}")
    logger.info(f"命名方式: {'使用命名内容' if config['naming_method'] == '1' else '使用自定义格式'}")
    if config['naming_method'] == '2':
        logger.info(f"前缀: {config['custom_prefix']}")
        logger.info(f"后缀: {config['custom_suffix']}")
        logger.info(f"连接符: {config['connector']}")
    logger.info("=" * 50)

def is_valid_image_url(url):
    """检查URL是否为有效的图片链接（使用更宽松的规则）"""
    if not url or not isinstance(url, str):
        return False
    
    url = url.strip().lower()
    
    # 检查是否是URL格式
    if not url.startswith(('http://', 'https://')):
        return False
    
    # 检查是否可能是图片URL（更宽松的规则）
    image_indicators = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '/image', '/img', 'images')
    return any(indicator in url for indicator in image_indicators)

def download_image(url, save_path, timeout=10, max_retries=3):
    """下载图片并保存到指定路径"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=timeout, stream=True)
            response.raise_for_status()
            
            # 检查内容类型（使用更宽松的规则）
            content_type = response.headers.get('content-type', '').lower()
            if not any(img_type in content_type for img_type in ['image', 'octet-stream', 'binary']):
                print(f"警告：内容类型可能不是图片 ({content_type})")
            
            # 保存图片
            with open(save_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            return True
            
        except requests.RequestException as e:
            if attempt == max_retries - 1:
                print(f"下载失败 ({url}): {str(e)}")
                return False
            time.sleep(1)  # 重试前等待
    
    return False

def detect_excel_type(ws, image_column, start_row):
    """检测Excel表格类型（图片/链接）"""
    check_rows = min(5, ws.max_row - start_row + 1)
    has_images = False
    has_urls = False
    
    try:
        # 使用 SheetImageLoader 检查图片
        image_loader = SheetImageLoader(ws)
        
        # 检查前几行是否包含图片
        for row in range(start_row, start_row + check_rows):
            cell_ref = f"{image_column}{row}"
            try:
                if image_loader.image_in(cell_ref):
                    has_images = True
                    break
            except:
                continue
        
        # 检查URL
        for row in range(start_row, start_row + check_rows):
            cell = ws[f"{image_column}{row}"]
            if cell.value and isinstance(cell.value, str):
                url = cell.value.strip()
                if url.startswith(('http://', 'https://')):
                    has_urls = True
                    break
                    
    except Exception as e:
        print_colored(f"类型检测警告: {str(e)}", Fore.YELLOW)
    
    # 返回检测结果
    if has_images and has_urls:
        return 'mixed'
    elif has_images:
        return 'images'
    elif has_urls:
        return 'urls'
    else:
        return 'unknown'

def print_colored(text, color=Fore.WHITE, style=Style.NORMAL, end='\n'):
    """打印彩色文本"""
    print(f"{style}{color}{text}{Style.RESET_ALL}", end=end)

def print_banner():
    """打印程序启动横幅"""
    banner = """
    ╔═══════════════════════════════════════════════╗
    ║             Excel 图片提取工具                ║
    ║    Excel Image Extractor v1.1 By Su007        ║
    ╚═══════════════════════════════════════════════╝
    """
    print_colored(banner, Fore.CYAN, Style.BRIGHT)

def print_header(text):
    """打印带样式的标题"""
    width = 60
    print("\n" + "─" * width)
    print_colored(f"┌{'─' * (width-2)}┐", Fore.CYAN, Style.BRIGHT)
    print_colored(f"│{text.center(width-2)}│", Fore.CYAN, Style.BRIGHT)
    print_colored(f"└{'─' * (width-2)}┘", Fore.CYAN, Style.BRIGHT)

def print_section(text):
    """打印带样式的分节标题"""
    print_colored(f"\n╭{'─' * (len(text) + 2)}╮", Fore.GREEN, Style.BRIGHT)
    print_colored(f"│ {text} │", Fore.GREEN, Style.BRIGHT)
    print_colored(f"╰{'─' * (len(text) + 2)}╯", Fore.GREEN)

def print_config_item(label, value, color=Fore.CYAN):
    """打印配置项"""
    print_colored(f"  ├─ {label}: ", color, end='')
    print_colored(value, Fore.WHITE)

def print_status(text, status="info"):
    """打印状态信息"""
    symbols = {
        "info": ("ℹ", Fore.BLUE),
        "success": ("✓", Fore.GREEN),
        "warning": ("⚠", Fore.YELLOW),
        "error": ("✗", Fore.RED)
    }
    symbol, color = symbols.get(status, symbols["info"])
    print_colored(f" {symbol} {text}", color)

def create_progress_bar(total, desc="处理进度"):
    """创建进度条"""
    return tqdm.tqdm(
        total=total,
        desc=desc,
        unit="张",
        ncols=80,
        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]"
    )

def convert_image_format(image, target_format):
    """转换图片格式，处理特殊情况"""
    if target_format.upper() in ['JPG', 'JPEG']:
        # 如果是JPG/JPEG格式，需要先转换为RGB模式
        if image.mode in ['RGBA', 'P']:
            # RGBA和P模式需要转换为RGB
            image = image.convert('RGB')
    return image

def save_image_with_format(image, file_path, image_format):
    """保存图片到指定格式，处理特殊情况"""
    try:
        # 转换图片格式
        converted_image = convert_image_format(image, image_format)
        
        # 统一处理JPG格式
        save_format = 'JPEG' if image_format.upper() in ['JPG', 'JPEG'] else image_format.upper()
        
        # 保存图片
        converted_image.save(file_path, format=save_format)
        return True
    except Exception as e:
        print_status(f"图片格式转换失败: {str(e)}", "error")
        return False

def main():
    # 检查是否有文件拖入
    if len(sys.argv) != 2:
        print_banner()
        print_status("请将Excel文件拖放到程序上运行！", "warning")
        input("\n按回车键退出...")
        return
        
    # 获取拖入的文件路径
    excel_path = Path(sys.argv[1])
    if not excel_path.exists() or not excel_path.suffix.lower() in ['.xlsx', '.xls']:
        print_banner()
        print_status("请拖入有效的Excel文件！", "error")
        input("\n按回车键退出...")
        return

    print_banner()
    
    try:
        # 预处理Excel文件，确保图片锚点信息完整
        print_status("正在预处理Excel文件...", "info")
        if not ensure_excel_anchors(excel_path):
            warning_msg = "文件预处理警告，尝试继续处理..."
            print_status(warning_msg, "warning")
            logger.warning(warning_msg)
        
        # 减少等待时间
        time.sleep(0.2)
        
        # 预先加载工作簿以获取工作表列表
        print_status("正在加载Excel文件...", "info")
        
        # 过滤特定的警告信息
        warnings.filterwarnings('ignore', category=UserWarning, 
                              module='openpyxl.reader.drawings')
        
        wb = load_workbook(excel_path)
        
        if not wb.sheetnames:
            print_status("错误：Excel文件中没有任何工作表！", "error")
            input("\n按回车键退出...")
            return
            
        # 获取用户配置
        CONFIG = get_config(excel_path, wb)
        
        # 检查工作表是否存在，如果不存在则提示重新选择
        while CONFIG['sheet_name'] not in wb.sheetnames:
            print_status(f"错误：工作表 '{CONFIG['sheet_name']}' 不存在！", "error")
            print_colored("\n当前Excel文件包含以下工作表：", Fore.YELLOW)
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                print_colored(f"  └─ {i}. {sheet_name}", Fore.CYAN)
            
            retry = input("\n是否重新选择工作表？(Y/N) [Y]: ").strip().upper()
            if retry != 'N':
                CONFIG = get_config(excel_path, wb, retry_sheet=True)
            else:
                print_status("程序终止", "error")
                input("\n按回车键退出...")
                return
        
        print_section("配置信息")
        print_config_item("工作表", CONFIG['sheet_name'])
        print_config_item("命名内容列", CONFIG['name_column'])
        print_config_item("图片列", CONFIG['image_column'])
        print_config_item("起始行", CONFIG['start_row'])
        print_config_item("图片格式", CONFIG['image_format'])
        print_config_item("命名方式", '使用命名内容' if CONFIG['naming_method'] == '1' else '使用自定义格式')
        
        if CONFIG['naming_method'] == '2':
            print_config_item("前缀", CONFIG['custom_prefix'] or '无')
            print_config_item("后缀", CONFIG['custom_suffix'] or '无')
            print_config_item("连接符", CONFIG['connector'] or '无')
            # 显示示例
            if not CONFIG['custom_prefix'] and not CONFIG['connector'] and CONFIG['custom_suffix']:
                print_status(f"文件将按数字顺序命名：1.{CONFIG['image_format']}、2.{CONFIG['image_format']}、3.{CONFIG['image_format']}...", "info")
            else:
                example_name = f"{CONFIG['custom_prefix']}{CONFIG['connector']}{CONFIG['custom_suffix']}"
                print_status(f"示例文件名: {example_name}.{CONFIG['image_format']}", "info")
        
        if input("\n确认配置是否正确？(Y/N) [Y]: ").strip().upper() != 'N':
            print_section("开始处理")
        else:
            print_status("\n重新配置...", "info")
            CONFIG = get_config(excel_path, wb)
            if CONFIG['sheet_name'] not in wb.sheetnames:
                print_status(f"错误：工作表 '{CONFIG['sheet_name']}' 不存在！", "error")
                return
        
        # 初始化路径
        base_path = Path(CONFIG['base_path'])
        save_dir = base_path / "图片"
        save_dir.mkdir(parents=True, exist_ok=True)
        excel_path = base_path / CONFIG['excel_filename']

        # 设置日志
        logger = setup_logging(save_dir)
        logger.info("开始处理Excel文件")
        log_config(logger, CONFIG)

        print_colored(f"工作目录: {base_path}", Fore.CYAN)
        print_colored(f"图片保存目录: {save_dir}", Fore.CYAN)
        print_colored(f"Excel文件: {excel_path}", Fore.CYAN)
        logger.info(f"工作目录: {base_path}")
        logger.info(f"图片保存目录: {save_dir}")
        logger.info(f"Excel文件: {excel_path}")

        # 加载工作表
        try:
            ws = wb[CONFIG['sheet_name']]
            
            # 检测Excel类型
            print_status("正在检测文件类型...", "info")
            excel_type = detect_excel_type(ws, CONFIG['image_column'], CONFIG['start_row'])
            print_colored(f"\n检测到Excel类型: {excel_type}", Fore.GREEN)
            logger.info(f"检测到Excel类型: {excel_type}")
            
            if excel_type == 'unknown':
                error_msg = "未检测到图片或有效的图片链接！"
                print_colored(f"\n{error_msg}", Fore.RED)
                logger.error(error_msg)
                print_colored("\n可能的原因：", Fore.YELLOW)
                print_colored("1. 图片未正确插入到Excel中", Fore.YELLOW)
                print_colored("2. 图片不在指定的列中", Fore.YELLOW)
                print_colored("3. 图片链接格式不正确", Fore.YELLOW)
                
                print_colored("\n建议操作：", Fore.CYAN)
                print_colored("1. 确认选择的列号是否正确", Fore.CYAN)
                print_colored("2. 检查图片是否正确插入到单元格中", Fore.CYAN)
                print_colored("3. 确保链接以http://或https://开头", Fore.CYAN)
                
                if input("\n是否重新选择列号？(Y/N) [Y]: ").strip().upper() != 'N':
                    CONFIG = get_config(excel_path, wb)
                    return
                else:
                    input("\n按回车键退出...")
                    return
            
            # 验证配置
            errors = validate_config(wb, CONFIG)
            if errors:
                logger.error("配置验证失败：")
                for error in errors:
                    logger.error(f"- {error}")
                print_colored("\n请重新配置...", Fore.RED)
                CONFIG = get_config(excel_path, wb)
                errors = validate_config(wb, CONFIG)
                if errors:
                    logger.error("配置仍然有误，程序终止")
                    print_colored("\n配置仍然有误，程序终止", Fore.RED)
                    input("按回车键退出...")
                    return
            
            # 初始化图片加载器
            image_loader = SheetImageLoader(ws)
            logger.info(f"成功加载工作表: {CONFIG['sheet_name']}")
            
            # 进度统计
            start_time = time.time()
            total = ws.max_row - CONFIG['start_row'] + 1
            processed = 0
            failed = 0
            
            # 创建进度条
            progress_bar = create_progress_bar(total)

            for row in range(CONFIG['start_row'], ws.max_row + 1):
                try:
                    name_cell = ws[f"{CONFIG['name_column']}{row}"]
                    
                    # 根据选择的命名方式生成文件名
                    if CONFIG['naming_method'] == '1':
                        if not name_cell.value:
                            progress_bar.write(f"跳过空行 {row}")
                            continue
                        clean_name = get_safe_filename(name_cell.value)
                    else:
                        if not CONFIG['custom_prefix'] and not CONFIG['connector'] and CONFIG['custom_suffix']:
                            clean_name = str(processed + 1)
                        else:
                            parts = []
                            if CONFIG['custom_prefix']:
                                parts.append(CONFIG['custom_prefix'])
                            if CONFIG['custom_suffix']:
                                parts.append(CONFIG['custom_suffix'])
                            clean_name = CONFIG['connector'].join(parts)

                    success = False
                    if excel_type in ['images', 'mixed']:
                        # 尝试提取嵌入图片
                        image = get_image_safely(image_loader, f"{CONFIG['image_column']}{row}")
                        if image:
                            try:
                                file_path = save_dir / f"{clean_name}.{CONFIG['image_format']}"
                                counter = 1
                                while file_path.exists():
                                    file_path = save_dir / f"{clean_name}_{counter}.{CONFIG['image_format']}"
                                    counter += 1
                                if save_image_with_format(image, file_path, CONFIG['image_format']):
                                    success = True
                            except Exception as e:
                                progress_bar.write(f"图片保存失败 [第{row}行]: {str(e)}")
                    
                    if not success and excel_type in ['urls', 'mixed']:
                        # 尝试下载链接图片
                        cell = ws[f"{CONFIG['image_column']}{row}"]
                        if cell.value and isinstance(cell.value, str):
                            url = cell.value.strip()
                            if url.startswith(('http://', 'https://')):
                                try:
                                    file_path = save_dir / f"{clean_name}.{CONFIG['image_format']}"
                                    counter = 1
                                    while file_path.exists():
                                        file_path = save_dir / f"{clean_name}_{counter}.{CONFIG['image_format']}"
                                        counter += 1
                                        
                                    # 先下载到临时文件
                                    temp_path = save_dir / f"temp_{clean_name}.{CONFIG['image_format']}"
                                    if download_image(url, temp_path):
                                        # 打开下载的图片并转换格式
                                        downloaded_image = Image.open(temp_path)
                                        if save_image_with_format(downloaded_image, file_path, CONFIG['image_format']):
                                            success = True
                                        downloaded_image.close()
                                        # 删除临时文件
                                        temp_path.unlink(missing_ok=True)
                                except Exception as e:
                                    progress_bar.write(f"链接处理失败 [第{row}行]: {str(e)}")
                                    # 确保删除临时文件
                                    if 'temp_path' in locals():
                                        temp_path.unlink(missing_ok=True)

                    if success:
                        processed += 1
                        logger.info(f"处理成功 [第{row}行]: {clean_name}")
                    else:
                        failed += 1
                        logger.error(f"处理失败 [第{row}行]")
                    
                    progress_bar.update(1)

                except Exception as e:
                    failed += 1
                    error_msg = f"错误 [第{row}行]: {str(e)}"
                    progress_bar.write(error_msg)
                    logger.error(error_msg)
                    progress_bar.update(1)

            progress_bar.close()
            
            # 完成统计
            end_time = time.time()
            duration = end_time - start_time
            
            print_section("处理完成")
            print_status(f"总行数: {total}行", "info")
            print_status(f"成功: {processed}个", "success")
            print_status(f"失败: {failed}个", "error" if failed > 0 else "info")
            print_status(f"耗时: {duration:.2f}秒", "info")
            logger.info("=" * 50)

        except Exception as e:
            error_msg = f"处理失败: {str(e)}"
            print_colored(error_msg, Fore.RED)
            logger.error(error_msg)
        
        print("\n处理完成！")
        logger.info("程序执行结束")
        input("按回车键退出...")

    except Exception as e:
        error_msg = f"处理失败: {str(e)}"
        print_colored(error_msg, Fore.RED)
        if 'logger' in locals():
            logger.error(error_msg)
    
    print("\n处理完成！")
    if 'logger' in locals():
        logger.info("程序执行结束")
    input("按回车键退出...")

if __name__ == "__main__":
    main()