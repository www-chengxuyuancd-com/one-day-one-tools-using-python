"""
Excel 图片提取工具

功能：
- 提取 Excel 中嵌入的图片（包括 EMF/WMF 等矢量格式自动转 PNG）
- 下载 Excel 单元格中图片链接对应的图片
- 支持按列提取（指定图片列 + 命名列）或提取全部嵌入图片
- 图片保存到与 Excel 同名的文件夹（可自选输出目录）
- 多种命名方式：顺序编号 / 前缀编号 / 链接文本 / 正则+起始数

UI 框架：PySide6
支持平台：Windows 10/11/11ARM, macOS Intel/Apple Silicon
打包方式：Nuitka
"""

import sys
import io
import os
import re
import time
import warnings
import zipfile
import subprocess
from pathlib import Path

# 添加项目根目录到路径，以便导入公共库
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from PySide6.QtWidgets import (
    QFileDialog, QComboBox, QLineEdit, QRadioButton, QButtonGroup,
    QHBoxLayout, QVBoxLayout, QFormLayout, QLabel, QWidget, QPushButton,
    QMessageBox, QGroupBox, QSpinBox, QStackedWidget,
)
from PySide6.QtCore import Qt

from PIL import Image
from openpyxl import load_workbook
import requests

from common.app_base import BaseApp
from common.utils import get_safe_filename

# 尝试导入图片加载器（用于按单元格位置提取图片）
try:
    from openpyxl_image_loader import SheetImageLoader
    HAS_IMAGE_LOADER = True
except ImportError:
    HAS_IMAGE_LOADER = False

# 过滤 openpyxl 的警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# EMF/WMF 常见的文件签名
_EMF_SIGNATURE = b'\x01\x00\x00\x00'  # EMF 文件头
_WMF_SIGNATURES = (b'\xd7\xcd\xc6\x9a', b'\x01\x00\x09\x00')  # WMF Aldus / 标准


# ================================================================
#  命名方式常量
# ================================================================
NAMING_SEQ = "seq"          # 顺序编号: 1, 2, 3
NAMING_PREFIX = "prefix"    # 前缀编号: Image_1, Image_2
NAMING_LINK = "link"        # 链接文本（按列模式专用）
NAMING_REGEX = "regex"      # 正则模板: img_{n}，支持起始数


class ExcelImageExtractor(BaseApp):
    """Excel 图片提取工具 - PySide6 GUI"""

    APP_NAME = "Excel 图片提取工具"
    APP_VERSION = "1.1"
    WINDOW_WIDTH = 920
    WINDOW_HEIGHT = 780

    IMAGE_FORMATS = ['png', 'jpg', 'jpeg', 'webp', 'bmp', 'gif']

    def __init__(self):
        self.wb = None
        self.excel_path = None
        super().__init__()

    # ================================================================
    #  UI 构建
    # ================================================================

    def create_content(self, layout):
        """创建配置界面"""
        form = QFormLayout()
        form.setSpacing(10)
        layout.addLayout(form)

        # ---- 文件选择 ----
        file_widget = QWidget()
        file_layout = QHBoxLayout(file_widget)
        file_layout.setContentsMargins(0, 0, 0, 0)

        self.file_input = QLineEdit()
        self.file_input.setReadOnly(True)
        self.file_input.setPlaceholderText("请选择 .xlsx 文件...")
        file_layout.addWidget(self.file_input, stretch=1)

        browse_btn = QPushButton("选择文件")
        browse_btn.setCursor(Qt.PointingHandCursor)
        browse_btn.clicked.connect(self._select_file)
        file_layout.addWidget(browse_btn)

        form.addRow("Excel 文件:", file_widget)

        # ---- 工作表选择 ----
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(260)
        form.addRow("工作表:", self.sheet_combo)

        # ---- 输出目录 ----
        outdir_widget = QWidget()
        outdir_layout = QHBoxLayout(outdir_widget)
        outdir_layout.setContentsMargins(0, 0, 0, 0)

        self.outdir_input = QLineEdit()
        self.outdir_input.setPlaceholderText("默认: 与 Excel 同名的文件夹")
        outdir_layout.addWidget(self.outdir_input, stretch=1)

        outdir_btn = QPushButton("选择目录")
        outdir_btn.setCursor(Qt.PointingHandCursor)
        outdir_btn.clicked.connect(self._select_output_dir)
        outdir_layout.addWidget(outdir_btn)

        form.addRow("输出目录:", outdir_widget)

        # ---- 提取模式 ----
        mode_widget = QWidget()
        mode_layout = QHBoxLayout(mode_widget)
        mode_layout.setContentsMargins(0, 0, 0, 0)

        self.mode_all = QRadioButton("提取全部嵌入图片")
        self.mode_all.setChecked(True)
        self.mode_col = QRadioButton("按列提取（嵌入图片 + 链接图片）")

        mode_layout.addWidget(self.mode_all)
        mode_layout.addWidget(self.mode_col)
        mode_layout.addStretch()

        form.addRow("提取模式:", mode_widget)

        # ---- 列配置区域（仅在「按列提取」模式下可见）----
        self.col_widget = QWidget()
        col_layout = QHBoxLayout(self.col_widget)
        col_layout.setContentsMargins(0, 0, 0, 0)
        col_layout.setSpacing(8)

        col_layout.addWidget(QLabel("图片列:"))
        self.img_col_input = QLineEdit("A")
        self.img_col_input.setMaximumWidth(60)
        col_layout.addWidget(self.img_col_input)

        col_layout.addSpacing(12)
        col_layout.addWidget(QLabel("命名列:"))
        self.name_col_input = QLineEdit("B")
        self.name_col_input.setMaximumWidth(60)
        col_layout.addWidget(self.name_col_input)

        col_layout.addSpacing(12)
        col_layout.addWidget(QLabel("起始行:"))
        self.start_row_input = QLineEdit("2")
        self.start_row_input.setMaximumWidth(60)
        col_layout.addWidget(self.start_row_input)

        hint = QLabel("（留空命名列则使用下方命名规则）")
        hint.setStyleSheet("color: #888; font-size: 11px;")
        col_layout.addSpacing(8)
        col_layout.addWidget(hint)
        col_layout.addStretch()

        self.col_widget.setVisible(False)
        form.addRow("", self.col_widget)

        # 模式切换 → 显示/隐藏列配置
        self.mode_col.toggled.connect(self.col_widget.setVisible)

        # ---- 图片格式 ----
        self.format_combo = QComboBox()
        self.format_combo.addItems(self.IMAGE_FORMATS)
        self.format_combo.setMaximumWidth(120)
        form.addRow("保存格式:", self.format_combo)

        # ---- 命名方式 ----
        naming_group = QGroupBox("命名方式")
        naming_outer = QVBoxLayout(naming_group)
        naming_outer.setSpacing(8)

        # 单选按钮行
        naming_btn_row = QHBoxLayout()
        self.naming_group = QButtonGroup(self)

        self.naming_seq_rb = QRadioButton("顺序编号 (1, 2, 3)")
        self.naming_seq_rb.setChecked(True)
        self.naming_prefix_rb = QRadioButton("前缀编号 (Image_1)")
        self.naming_link_rb = QRadioButton("链接文本")
        self.naming_regex_rb = QRadioButton("自定义模板")

        self.naming_group.addButton(self.naming_seq_rb)
        self.naming_group.addButton(self.naming_prefix_rb)
        self.naming_group.addButton(self.naming_link_rb)
        self.naming_group.addButton(self.naming_regex_rb)

        naming_btn_row.addWidget(self.naming_seq_rb)
        naming_btn_row.addWidget(self.naming_prefix_rb)
        naming_btn_row.addWidget(self.naming_link_rb)
        naming_btn_row.addWidget(self.naming_regex_rb)
        naming_btn_row.addStretch()
        naming_outer.addLayout(naming_btn_row)

        # 各命名方式的配置面板（用 QStackedWidget 切换）
        self.naming_stack = QStackedWidget()

        # page 0: 顺序编号 - 起始数
        seq_page = QWidget()
        seq_layout = QHBoxLayout(seq_page)
        seq_layout.setContentsMargins(0, 0, 0, 0)
        seq_layout.addWidget(QLabel("起始数:"))
        self.seq_start_spin = QSpinBox()
        self.seq_start_spin.setRange(0, 999999)
        self.seq_start_spin.setValue(1)
        self.seq_start_spin.setMaximumWidth(100)
        seq_layout.addWidget(self.seq_start_spin)
        seq_layout.addStretch()
        self.naming_stack.addWidget(seq_page)

        # page 1: 前缀编号
        prefix_page = QWidget()
        prefix_layout = QHBoxLayout(prefix_page)
        prefix_layout.setContentsMargins(0, 0, 0, 0)
        prefix_layout.addWidget(QLabel("前缀:"))
        self.prefix_input = QLineEdit("Image")
        self.prefix_input.setMaximumWidth(120)
        prefix_layout.addWidget(self.prefix_input)
        prefix_layout.addWidget(QLabel("连接符:"))
        self.prefix_sep_input = QLineEdit("_")
        self.prefix_sep_input.setMaximumWidth(40)
        prefix_layout.addWidget(self.prefix_sep_input)
        prefix_layout.addWidget(QLabel("起始数:"))
        self.prefix_start_spin = QSpinBox()
        self.prefix_start_spin.setRange(0, 999999)
        self.prefix_start_spin.setValue(1)
        self.prefix_start_spin.setMaximumWidth(100)
        prefix_layout.addWidget(self.prefix_start_spin)
        prefix_lbl = QLabel("预览: Image_1")
        prefix_lbl.setStyleSheet("color: #888;")
        prefix_layout.addSpacing(10)
        prefix_layout.addWidget(prefix_lbl)
        self._prefix_preview_lbl = prefix_lbl
        prefix_layout.addStretch()
        self.naming_stack.addWidget(prefix_page)

        # 前缀预览联动
        self.prefix_input.textChanged.connect(self._update_prefix_preview)
        self.prefix_sep_input.textChanged.connect(self._update_prefix_preview)
        self.prefix_start_spin.valueChanged.connect(self._update_prefix_preview)

        # page 2: 链接文本
        link_page = QWidget()
        link_layout = QHBoxLayout(link_page)
        link_layout.setContentsMargins(0, 0, 0, 0)
        link_hint = QLabel("使用链接的单元格文本作为文件名（仅在按列提取模式下有效，无链接文本时自动回退编号）")
        link_hint.setStyleSheet("color: #888; font-size: 11px;")
        link_layout.addWidget(link_hint)
        link_layout.addStretch()
        self.naming_stack.addWidget(link_page)

        # page 3: 自定义模板
        regex_page = QWidget()
        regex_layout = QHBoxLayout(regex_page)
        regex_layout.setContentsMargins(0, 0, 0, 0)
        regex_layout.addWidget(QLabel("模板:"))
        self.regex_input = QLineEdit("img_{n}")
        self.regex_input.setMaximumWidth(200)
        regex_layout.addWidget(self.regex_input)
        regex_layout.addWidget(QLabel("起始数:"))
        self.regex_start_spin = QSpinBox()
        self.regex_start_spin.setRange(0, 999999)
        self.regex_start_spin.setValue(1)
        self.regex_start_spin.setMaximumWidth(100)
        regex_layout.addWidget(self.regex_start_spin)
        regex_hint = QLabel("{n} 表示编号")
        regex_hint.setStyleSheet("color: #888; font-size: 11px;")
        regex_layout.addWidget(regex_hint)
        regex_layout.addStretch()
        self.naming_stack.addWidget(regex_page)

        naming_outer.addWidget(self.naming_stack)
        layout.addWidget(naming_group)

        # 切换命名方式 → 切换配置面板
        self.naming_seq_rb.toggled.connect(
            lambda c: c and self.naming_stack.setCurrentIndex(0)
        )
        self.naming_prefix_rb.toggled.connect(
            lambda c: c and self.naming_stack.setCurrentIndex(1)
        )
        self.naming_link_rb.toggled.connect(
            lambda c: c and self.naming_stack.setCurrentIndex(2)
        )
        self.naming_regex_rb.toggled.connect(
            lambda c: c and self.naming_stack.setCurrentIndex(3)
        )

    def _update_prefix_preview(self):
        """前缀编号预览"""
        prefix = self.prefix_input.text()
        sep = self.prefix_sep_input.text()
        start = self.prefix_start_spin.value()
        self._prefix_preview_lbl.setText(f"预览: {prefix}{sep}{start}")

    # ================================================================
    #  文件操作
    # ================================================================

    def _select_file(self):
        """选择 Excel 文件"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if filepath:
            self.file_input.setText(filepath)
            self.excel_path = Path(filepath)
            # 默认输出目录 = 同名文件夹
            default_out = str(self.excel_path.parent / self.excel_path.stem)
            self.outdir_input.setPlaceholderText(f"默认: {default_out}")
            self._load_workbook()

    def _select_output_dir(self):
        """选择自定义输出目录"""
        start_dir = self.outdir_input.text() or (
            str(self.excel_path.parent) if self.excel_path else ""
        )
        dirpath = QFileDialog.getExistingDirectory(
            self, "选择输出目录", start_dir
        )
        if dirpath:
            self.outdir_input.setText(dirpath)

    def _get_output_dir(self):
        """获取最终输出目录（用户自选 > 默认同名文件夹）"""
        custom = self.outdir_input.text().strip()
        if custom:
            return Path(custom)
        return self.excel_path.parent / self.excel_path.stem

    def _load_workbook(self):
        """加载工作簿并填充工作表下拉列表"""
        try:
            if self.wb:
                try:
                    self.wb.close()
                except Exception:
                    pass
            self.wb = load_workbook(self.excel_path, data_only=True)
            sheets = self.wb.sheetnames
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            if sheets:
                self.sheet_combo.setCurrentIndex(0)
            self.log(
                f"已加载: {self.excel_path.name}（{len(sheets)} 个工作表）",
                "success"
            )
        except Exception as e:
            self.log(f"加载文件失败: {e}", "error")
            QMessageBox.critical(self, "错误", f"无法加载 Excel 文件:\n{e}")

    # ================================================================
    #  命名逻辑
    # ================================================================

    def _get_naming_mode(self):
        """获取当前选择的命名方式"""
        if self.naming_seq_rb.isChecked():
            return NAMING_SEQ
        elif self.naming_prefix_rb.isChecked():
            return NAMING_PREFIX
        elif self.naming_link_rb.isChecked():
            return NAMING_LINK
        else:
            return NAMING_REGEX

    def _make_name(self, mode, counter, link_text=None):
        """
        根据命名模式和计数器生成文件名

        :param mode: 命名模式常量
        :param counter: 当前顺序计数器值（已经加上了起始偏移）
        :param link_text: 链接文本（仅 NAMING_LINK 模式使用）
        :return: 安全文件名（不含扩展名）
        """
        if mode == NAMING_SEQ:
            return str(counter)

        elif mode == NAMING_PREFIX:
            prefix = self.prefix_input.text() or "Image"
            sep = self.prefix_sep_input.text()
            return f"{prefix}{sep}{counter}"

        elif mode == NAMING_LINK:
            if link_text and str(link_text).strip():
                return get_safe_filename(str(link_text))
            # 回退到顺序编号
            return str(counter)

        elif mode == NAMING_REGEX:
            tpl = self.regex_input.text() or "img_{n}"
            return tpl.replace("{n}", str(counter))

        return str(counter)

    def _get_start_number(self, mode):
        """获取当前命名模式的起始编号"""
        if mode == NAMING_SEQ:
            return self.seq_start_spin.value()
        elif mode == NAMING_PREFIX:
            return self.prefix_start_spin.value()
        elif mode == NAMING_REGEX:
            return self.regex_start_spin.value()
        return 1

    # ================================================================
    #  校验
    # ================================================================

    def validate(self):
        """开始前校验"""
        if not self.excel_path or not self.excel_path.exists():
            QMessageBox.warning(self, "提示", "请先选择 Excel 文件")
            return False

        if self.excel_path.suffix.lower() != '.xlsx':
            QMessageBox.warning(self, "提示", "仅支持 .xlsx 格式的 Excel 文件")
            return False

        if not self.sheet_combo.currentText():
            QMessageBox.warning(self, "提示", "请选择工作表")
            return False

        if self.mode_col.isChecked():
            img_col = self.img_col_input.text().strip().upper()
            if not img_col or not re.match(r'^[A-Z]+$', img_col):
                QMessageBox.warning(
                    self, "提示", "请输入有效的图片列（如 A、B、AA）"
                )
                return False

            name_col = self.name_col_input.text().strip().upper()
            if name_col and not re.match(r'^[A-Z]+$', name_col):
                QMessageBox.warning(
                    self, "提示",
                    "命名列格式无效（如 A、B、AA），留空则使用命名规则"
                )
                return False

            start_row = self.start_row_input.text().strip()
            if not start_row.isdigit() or int(start_row) < 1:
                QMessageBox.warning(self, "提示", "起始行必须是大于 0 的数字")
                return False

        return True

    # ================================================================
    #  核心提取逻辑
    # ================================================================

    def do_work(self):
        """执行图片提取（在后台线程中运行）"""
        output_dir = self._get_output_dir()
        output_dir.mkdir(parents=True, exist_ok=True)

        self.log(f"输出目录: {output_dir}")

        sheet_name = self.sheet_combo.currentText()
        image_format = self.format_combo.currentText()

        try:
            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb[sheet_name]
            self.log(f"已加载工作表: {sheet_name}")
        except Exception as e:
            self.log(f"加载工作表失败: {e}", "error")
            return

        try:
            if self.mode_all.isChecked():
                self._extract_all_images(ws, output_dir, image_format)
            else:
                self._extract_by_column(ws, output_dir, image_format)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # ----------------------------------------------------------------

    def _extract_all_images(self, ws, output_dir, image_format):
        """模式一：从 xlsx 中提取所有嵌入图片（通过 zip 解压媒体文件）"""
        self.log("模式: 提取全部嵌入图片")

        media_files = []
        try:
            with zipfile.ZipFile(self.excel_path, 'r') as zf:
                for name in zf.namelist():
                    # xl/media/ 下的所有文件，排除目录本身
                    if name.startswith('xl/media/') and not name.endswith('/'):
                        media_files.append(name)
        except Exception as e:
            self.log(f"读取文件内部结构失败: {e}", "error")
            return

        # 按文件名中的数字排序
        def _sort_key(fname):
            nums = re.findall(r'(\d+)', fname)
            return int(nums[-1]) if nums else 0

        media_files.sort(key=_sort_key)
        total = len(media_files)

        if total == 0:
            self.log("未在文件中找到嵌入图片", "warning")
            self.update_progress(100, "完成 - 未找到图片")
            return

        self.log(f"找到 {total} 个媒体文件，开始提取...")
        naming_mode = self._get_naming_mode()
        start_num = self._get_start_number(naming_mode)

        success = 0
        failed = 0
        counter = start_num

        with zipfile.ZipFile(self.excel_path, 'r') as zf:
            for i, media_name in enumerate(media_files):
                if self.should_stop:
                    self.log("用户取消操作", "warning")
                    break

                try:
                    image_data = zf.read(media_name)

                    # 尝试用 Pillow 打开
                    pil_image = self._open_image_data(
                        image_data, media_name, output_dir
                    )
                    if pil_image is None:
                        failed += 1
                        self.log(
                            f"[{i + 1}/{total}] 不支持的格式，已跳过 ({media_name})",
                            "warning"
                        )
                        self._update_extract_progress(
                            i, total, success, failed
                        )
                        continue

                    filename = self._make_name(naming_mode, counter)
                    filepath = self._get_unique_path(
                        output_dir, filename, image_format
                    )
                    self._save_image(pil_image, filepath, image_format)
                    success += 1
                    counter += 1
                    self.log(
                        f"[{i + 1}/{total}] 保存: {filepath.name}", "success"
                    )

                except Exception as e:
                    failed += 1
                    self.log(
                        f"[{i + 1}/{total}] 提取失败 ({media_name}): {e}",
                        "error"
                    )

                self._update_extract_progress(i, total, success, failed)

        self._print_summary(total, success, failed, output_dir)

    # ----------------------------------------------------------------

    def _extract_by_column(self, ws, output_dir, image_format):
        """模式二：按列提取图片（嵌入图片 + URL 图片）"""
        img_col = self.img_col_input.text().strip().upper()
        name_col = self.name_col_input.text().strip().upper()
        start_row = int(self.start_row_input.text().strip())
        naming_mode = self._get_naming_mode()
        start_num = self._get_start_number(naming_mode)

        self.log("模式: 按列提取")
        self.log(
            f"  图片列: {img_col} | 命名列: {name_col or '(使用命名规则)'} "
            f"| 起始行: {start_row}"
        )

        # 初始化嵌入图片加载器
        image_loader = None
        if HAS_IMAGE_LOADER:
            try:
                image_loader = SheetImageLoader(ws)
                self.log("嵌入图片加载器已就绪")
            except Exception as e:
                self.log(
                    f"嵌入图片加载器初始化失败（将只处理链接图片）: {e}",
                    "warning"
                )
        else:
            self.log("未安装 openpyxl-image-loader，将只处理链接图片", "warning")
            self.log("  安装方式: pip install openpyxl-image-loader", "info")

        max_row = ws.max_row
        if max_row is None or max_row < start_row:
            self.log("没有数据行可处理", "warning")
            self.update_progress(100, "完成 - 无数据")
            return

        total = max_row - start_row + 1
        self.log(f"准备处理 {total} 行数据（行 {start_row} ~ {max_row}）")

        success = 0
        failed = 0
        skipped = 0
        counter = start_num

        for row_idx in range(start_row, max_row + 1):
            if self.should_stop:
                self.log("用户取消操作", "warning")
                break

            current = row_idx - start_row + 1

            try:
                # 确定文件命名
                clean_name = self._resolve_column_name(
                    ws, row_idx, name_col, img_col, naming_mode, counter
                )

                saved = False

                # --- 尝试 1: 提取嵌入图片 ---
                if image_loader:
                    cell_ref = f"{img_col}{row_idx}"
                    try:
                        if image_loader.image_in(cell_ref):
                            pil_image = image_loader.get(cell_ref)
                            filepath = self._get_unique_path(
                                output_dir, clean_name, image_format
                            )
                            self._save_image(pil_image, filepath, image_format)
                            success += 1
                            counter += 1
                            saved = True
                            self.log(
                                f"[行{row_idx}] 嵌入图片 → {filepath.name}",
                                "success"
                            )
                    except Exception as e:
                        self.log(
                            f"[行{row_idx}] 嵌入图片提取失败: {e}", "warning"
                        )

                # --- 尝试 2: 下载链接图片 ---
                if not saved:
                    cell = ws[f"{img_col}{row_idx}"]
                    url = self._get_url_from_cell(cell)

                    if url:
                        filepath = self._get_unique_path(
                            output_dir, clean_name, image_format
                        )
                        if self._download_and_save(url, filepath, image_format):
                            success += 1
                            counter += 1
                            saved = True
                            self.log(
                                f"[行{row_idx}] 链接图片 → {filepath.name}",
                                "success"
                            )
                        else:
                            failed += 1
                            self.log(
                                f"[行{row_idx}] 下载失败: {url[:80]}...",
                                "error"
                            )

                if not saved:
                    skipped += 1

            except Exception as e:
                failed += 1
                self.log(f"[行{row_idx}] 处理出错: {e}", "error")

            self.update_progress(
                current / total * 100,
                f"行 {row_idx}/{max_row} | "
                f"成功: {success} | 失败: {failed} | 跳过: {skipped}"
            )

        self._print_summary(total, success, failed, output_dir, skipped)

    # ================================================================
    #  辅助方法
    # ================================================================

    def _resolve_column_name(
        self, ws, row_idx, name_col, img_col, naming_mode, counter
    ):
        """按列模式下解析文件命名"""
        # 如果指定了命名列且该列有值，优先使用（但排除 URL 值）
        if name_col:
            cell_value = ws[f"{name_col}{row_idx}"].value
            if cell_value is not None:
                text = str(cell_value).strip()
                # 如果值是 URL，不作为文件名使用
                if text and not text.startswith(('http://', 'https://')):
                    return get_safe_filename(text)

        # 链接文本模式：尝试获取单元格的显示文本（非 URL 部分）
        if naming_mode == NAMING_LINK:
            cell = ws[f"{img_col}{row_idx}"]
            link_text = cell.value
            # 使用超链接的显示文本（不是 URL 本身）
            if link_text and str(link_text).strip():
                display = str(link_text).strip()
                if not display.startswith(('http://', 'https://')):
                    return get_safe_filename(display)
            # 也尝试超链接的 tooltip / display
            if cell.hyperlink:
                hl = cell.hyperlink
                # 有些超链接有 display 属性
                if hasattr(hl, 'display') and hl.display:
                    disp = str(hl.display).strip()
                    if disp and not disp.startswith(('http://', 'https://')):
                        return get_safe_filename(disp)

        # 其他模式 / 回退：用命名规则生成
        return self._make_name(naming_mode, counter)

    @staticmethod
    def _open_image_data(data, media_name, output_dir):
        """
        尝试将二进制数据打开为 PIL Image。
        对于 EMF/WMF 等 Pillow 不支持的格式，尝试转换。
        返回 PIL Image 或 None。
        """
        # 先直接尝试 Pillow 打开
        try:
            return Image.open(io.BytesIO(data))
        except Exception:
            pass

        # 检测是否是 EMF/WMF
        ext = Path(media_name).suffix.lower()
        is_emf_wmf = ext in ('.emf', '.wmf')

        # 也通过文件头签名检测
        if not is_emf_wmf and len(data) >= 4:
            header = data[:4]
            if header == _EMF_SIGNATURE or header in _WMF_SIGNATURES:
                is_emf_wmf = True

        if not is_emf_wmf:
            return None

        # 尝试用 sips (macOS) 或 magick (ImageMagick) 转换 EMF/WMF → PNG
        return _convert_vector_to_image(data, ext, output_dir)

    @staticmethod
    def _get_url_from_cell(cell):
        """从单元格中提取 URL（支持单元格值和超链接）"""
        if cell.hyperlink and cell.hyperlink.target:
            url = cell.hyperlink.target.strip()
            if url.startswith(('http://', 'https://')):
                return url

        if cell.value and isinstance(cell.value, str):
            url = cell.value.strip()
            if url.startswith(('http://', 'https://')):
                return url

        return None

    @staticmethod
    def _get_unique_path(directory, name, fmt):
        """生成不重复的文件路径"""
        filepath = directory / f"{name}.{fmt}"
        counter = 1
        while filepath.exists():
            filepath = directory / f"{name}_{counter}.{fmt}"
            counter += 1
        return filepath

    @staticmethod
    def _save_image(pil_image, filepath, image_format):
        """保存 PIL 图片到指定格式"""
        save_fmt = (
            'JPEG' if image_format.lower() in ('jpg', 'jpeg')
            else image_format.upper()
        )

        if save_fmt == 'JPEG' and pil_image.mode in ('RGBA', 'P', 'LA', 'PA'):
            pil_image = pil_image.convert('RGB')
        elif pil_image.mode == 'P' and save_fmt == 'PNG':
            pil_image = pil_image.convert('RGBA')

        pil_image.save(filepath, format=save_fmt)

    def _download_and_save(
        self, url, filepath, image_format, timeout=15, max_retries=3
    ):
        """下载图片并保存到指定路径，支持自动重试"""
        headers = {
            'User-Agent': (
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                'AppleWebKit/537.36 (KHTML, like Gecko) '
                'Chrome/120.0.0.0 Safari/537.36'
            )
        }

        for attempt in range(max_retries):
            if self.should_stop:
                return False
            try:
                resp = requests.get(
                    url, headers=headers, timeout=timeout, stream=True
                )
                resp.raise_for_status()

                image_data = io.BytesIO(resp.content)
                pil_image = Image.open(image_data)
                self._save_image(pil_image, filepath, image_format)
                return True

            except Exception as e:
                if attempt < max_retries - 1:
                    self.log(
                        f"  下载重试 ({attempt + 1}/{max_retries}): {e}",
                        "warning"
                    )
                    time.sleep(1 * (attempt + 1))
                else:
                    self.log(f"  下载最终失败: {e}", "error")

        return False

    def _update_extract_progress(self, i, total, success, failed):
        """更新提取进度"""
        self.update_progress(
            (i + 1) / total * 100,
            f"进度: {i + 1}/{total} | 成功: {success} | 失败: {failed}"
        )

    def _print_summary(self, total, success, failed, output_dir, skipped=0):
        """输出处理结果摘要"""
        self.log("=" * 50)
        self.log("处理完成!")
        self.log(f"  总计: {total}")
        self.log(f"  成功: {success}", "success" if success > 0 else "info")
        self.log(f"  失败: {failed}", "error" if failed > 0 else "info")
        if skipped > 0:
            self.log(f"  跳过: {skipped}")
        self.log(f"  输出目录: {output_dir}", "info")
        self.update_progress(100, f"完成 - 成功: {success} | 失败: {failed}")


# ================================================================
#  EMF/WMF 转换辅助
# ================================================================

def _convert_vector_to_image(data, ext, output_dir):
    """
    尝试将 EMF/WMF 数据转为 PNG（利用系统工具）。
    - macOS: 使用 sips
    - 其他: 尝试 ImageMagick (magick/convert)
    返回 PIL Image 或 None
    """
    import tempfile

    suffix = ext if ext.startswith('.') else f'.{ext}'
    tmp_in = None
    tmp_out = None

    try:
        # 写入临时源文件
        tmp_in = tempfile.NamedTemporaryFile(
            suffix=suffix, delete=False, dir=str(output_dir)
        )
        tmp_in.write(data)
        tmp_in.close()

        tmp_out_path = tmp_in.name + '.png'

        converted = False

        if sys.platform == 'darwin':
            # macOS sips 可以处理部分 EMF
            try:
                subprocess.run(
                    ['sips', '-s', 'format', 'png', tmp_in.name,
                     '--out', tmp_out_path],
                    capture_output=True, timeout=10, check=True
                )
                converted = True
            except Exception:
                pass

        # 尝试 ImageMagick
        if not converted:
            for cmd in ('magick', 'convert'):
                try:
                    subprocess.run(
                        [cmd, tmp_in.name, tmp_out_path],
                        capture_output=True, timeout=15, check=True
                    )
                    converted = True
                    break
                except Exception:
                    continue

        if converted and os.path.exists(tmp_out_path):
            img = Image.open(tmp_out_path)
            img.load()  # 确保数据已读入内存
            return img

    except Exception:
        pass
    finally:
        # 清理临时文件
        for p in (tmp_in.name if tmp_in else None, tmp_out_path if 'tmp_out_path' in dir() else None):
            if p:
                try:
                    os.unlink(p)
                except Exception:
                    pass

    return None


# ================================================================
#  入口
# ================================================================

if __name__ == "__main__":
    app = ExcelImageExtractor()
    app.run()
